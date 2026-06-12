import os
import uuid
import json
import hashlib
import secrets
import requests
from datetime import datetime
from flask import Flask, jsonify, request, session, redirect
from flask_cors import CORS
from supabase import create_client, Client

app = Flask(__name__)
app.secret_key = 'secret_key_for_session_12345'

# === CORS ===
CORS(app, 
     origins=["https://1lyas1k333.github.io", "https://arturchik-box-2.onrender.com", "http://127.0.0.1:5500", "http://localhost:5500"],
     supports_credentials=True,
     allow_headers=["Content-Type", "Authorization"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])

@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = app.make_default_options_response()
        response.headers.add('Access-Control-Allow-Origin', 'https://1lyas1k333.github.io')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        response.headers.add('Access-Control-Allow-Credentials', 'true')
        return response

# === SUPABASE ===
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://xszdufdzgvzwtiyppyxs.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# === TELEGRAM ===
TELEGRAM_TOKEN = "8694164916:AAEYQey-DSovguWmgy-mZLG4nMVhSV4BunQ"
TELEGRAM_CHAT_ID = "1056646376"

def send_telegram_message(message):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        requests.post(url, json={"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"}, timeout=5)
        print(f"[TG] Уведомление отправлено админу")
    except Exception as e:
        print(f"[TG] Ошибка: {e}")

def send_telegram_to_user(chat_id, message):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        requests.post(url, json={"chat_id": chat_id, "text": message, "parse_mode": "HTML"}, timeout=5)
        print(f"[TG_USER] Отправлено пользователю {chat_id}")
    except Exception as e:
        print(f"[TG_USER] Ошибка: {e}")

# === PLATEGA ===
PLATEGA_SHOP_ID = "a8922d02-2beb-44a0-b24a-4b6e6caa33ef"
PLATEGA_API_KEY = "osj9xJrzJb9jeFXjUHBMucfuR8DXydxScLOGImdzGaiMXNLj8KuiBDsH3AUBZ1vlsckfPWD4jZhdw5HQzJPJdQJWTkitFDtBCAtL"
PLATEGA_API_URL = "https://app.platega.io/transaction/process"

# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def generate_token(user_id):
    data = f"{user_id}:{datetime.now().isoformat()}:{secrets.token_hex(8)}"
    return hashlib.sha256(data.encode()).hexdigest()

def update_order_status(order_id, status):
    # Обновляем статус в БД
    supabase.table("orders").update({"status": status, "updated_at": datetime.now().isoformat()}).eq("order_id", order_id).execute()
    
    # Уведомление админу
    status_text = {
        'pending': '⏳ Ожидает оплаты',
        'paid': '✅ Оплачен',
        'shipped': '📦 Отправлен',
        'completed': '🎉 Завершён'
    }.get(status, status)
    
    admin_msg = f"""🔄 <b>СТАТУС ЗАКАЗА ИЗМЕНЁН</b>
📦 Заказ: {order_id}
📌 Новый статус: {status_text}
🔗 Админка: https://arturchik-box-2.onrender.com/admin"""
    send_telegram_message(admin_msg)
    
    # 👇 ДОБАВЛЯЕМ УВЕДОМЛЕНИЕ КЛИЕНТУ ОБ ОПЛАТЕ
    if status == 'paid':
        res = supabase.table("orders").select("customer_name, user_id").eq("order_id", order_id).execute()
        if res.data:
            customer_name = res.data[0]['customer_name']
            user_id = res.data[0]['user_id']
            user_res = supabase.table("users").select("telegram_id").eq("id", user_id).execute()
            if user_res.data and user_res.data[0].get('telegram_id'):
                client_msg = f"""📦 <b>АРТУРЧИК box</b>
            
Здравствуйте, {customer_name}!
Статус вашего заказа <b>№{order_id}</b> изменился на:
✅ Оплачен

Спасибо, что выбрали нас!"""
                send_telegram_to_user(user_res.data[0]['telegram_id'], client_msg)
    
    print(f"[DB] Заказ {order_id} обновлён → {status}")

# === API ===
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    user_id = str(uuid.uuid4())
    hashed_pw = hash_password(data['password'])
    try:
        supabase.table("users").insert({
            "id": user_id, "email": data['email'], "password": hashed_pw,
            "name": data['name'], "phone": data['phone'], "telegram_id": data.get('telegram'),
            "created_at": datetime.now().isoformat()
        }).execute()
        return jsonify({"success": True, "user_id": user_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    hashed_input = hash_password(data['password'])
    res = supabase.table("users").select("*").eq("email", data['email']).eq("password", hashed_input).execute()
    if not res.data:
        return jsonify({"success": False, "error": "Неверный email или пароль"}), 401
    user = res.data[0]
    token = generate_token(user['id'])
    supabase.table("user_tokens").insert({
        "id": str(uuid.uuid4()), "user_id": user['id'], "token": token,
        "created_at": datetime.now().isoformat(), "expires_at": datetime.now().timestamp() + 30*24*3600
    }).execute()
    return jsonify({"success": True, "token": token, "user": {"id": user['id'], "name": user['name'], "email": user['email']}})

@app.route('/api/my-orders', methods=['GET'])
def get_my_orders():
    auth = request.headers.get('Authorization')
    if not auth or not auth.startswith('Bearer '):
        return jsonify({"success": False, "error": "Не авторизован"}), 401
    token = auth.split(' ')[1]
    res = supabase.table("user_tokens").select("user_id").eq("token", token).execute()
    if not res.data:
        return jsonify({"success": False, "error": "Не авторизован"}), 401
    user_id = res.data[0]['user_id']
    orders = supabase.table("orders").select("*").eq("user_id", user_id).execute()
    return jsonify({"success": True, "orders": orders.data})

@app.route('/api/orders', methods=['GET'])
def get_all_orders():
    orders = supabase.table("orders").select("*").execute()
    return jsonify({"success": True, "orders": orders.data, "count": len(orders.data)})

@app.route('/api/update-status', methods=['POST'])
def update_status_api():
    data = request.get_json()
    order_id = data.get('order_id')
    status = data.get('status')
    if not order_id or not status:
        return jsonify({"success": False, "error": "Missing fields"}), 400
    update_order_status(order_id, status)
    return jsonify({"success": True})

@app.route('/api/update-tracking', methods=['POST'])
def update_tracking():
    data = request.get_json()
    order_id = data.get('order_id')
    tracking_number = data.get('tracking_number')
    if not order_id:
        return jsonify({"success": False, "error": "order_id required"}), 400
    
    supabase.table("orders").update({"tracking_number": tracking_number}).eq("order_id", order_id).execute()
    
    # Уведомление покупателю
    res = supabase.table("orders").select("customer_name, user_id").eq("order_id", order_id).execute()
    if res.data:
        customer_name = res.data[0]['customer_name']
        user_id = res.data[0]['user_id']
        user_res = supabase.table("users").select("telegram_id").eq("id", user_id).execute()
        if user_res.data and user_res.data[0].get('telegram_id'):
            msg_user = f"""📦 <b>АРТУРЧИК box</b>

Здравствуйте, {customer_name}!

Ваш заказ <b>№{order_id}</b> отправлен!

📦 Трек-номер: {tracking_number}
🔗 Отследить: https://www.cdek.ru/track?order_id={tracking_number}

Спасибо за покупку!"""
            send_telegram_to_user(user_res.data[0]['telegram_id'], msg_user)
    
    return jsonify({"success": True})

@app.route('/create-payment', methods=['POST'])
def create_payment():
    data = request.get_json()
    amount = data.get('amount', 3490)
    cart_items = data.get('items', [])
    customer = data.get('customer', {})
    user_id = data.get('user_id')
    
    customer_name = customer.get('fullName', '')
    customer_phone = customer.get('phone', '')
    customer_email = customer.get('email', '')
    customer_address = customer.get('address', '')
    customer_city = customer.get('city', '')
    customer_extra = customer.get('extraAddress', '')
    customer_notes = customer.get('notes', '')
    
    order_id = f"ORDER_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    
    # Сохраняем заказ
    order_data = {
        "id": str(uuid.uuid4()), "order_id": order_id, "user_id": user_id,
        "customer_name": customer_name, "customer_phone": customer_phone,
        "customer_email": customer_email, "customer_address": customer_address,
        "customer_city": customer_city, "customer_extra": customer_extra,
        "customer_notes": customer_notes, "items": json.dumps(cart_items),
        "total_amount": amount, "status": "pending", 
        "created_at": datetime.now().isoformat(), "updated_at": datetime.now().isoformat()
    }
    supabase.table("orders").insert(order_data).execute()
    
    # Уведомление админу о новом заказе
    admin_msg = f"""🆕 <b>НОВЫЙ ЗАКАЗ!</b>

📦 Заказ: {order_id}
👤 Клиент: {customer_name}
📞 Телефон: {customer_phone}
📧 Email: {customer_email}
💰 Сумма: {amount} ₽
📍 Адрес: {customer_city}, {customer_address}

🔗 Админка: https://arturchik-box-2.onrender.com/admin"""
    send_telegram_message(admin_msg)

    # Уведомление клиенту о создании заказа (если привязан Telegram)
    if user_id:
        user_res = supabase.table("users").select("telegram_id").eq("id", user_id).execute()
        if user_res.data and user_res.data[0].get('telegram_id'):
            items_text = ', '.join([f"{item.get('name', '')} ({item.get('size', '')}) x{item.get('quantity', 1)}" for item in cart_items])
            client_msg = f"""📦 <b>АРТУРЧИК box</b>
                
Здравствуйте, {customer_name}!
Ваш заказ <b>№{order_id}</b> успешно создан.

📋 Состав заказа:
{items_text}

💰 Сумма: {amount} ₽
📌 Статус: ⏳ Ожидает оплаты

✅ После оплаты мы отправим вам трек-номер для отслеживания.

Спасибо за покупку!"""
            send_telegram_to_user(user_res.data[0]['telegram_id'], client_msg)

    # Создаём платёж в Platega
    headers = {
        "X-MerchantId": PLATEGA_SHOP_ID,
        "X-Secret": PLATEGA_API_KEY,
        "Content-Type": "application/json"
    }
    payload = {
    "paymentMethod": 2,
    "paymentDetails": {"amount": float(amount), "currency": "RUB"},
    "description": f"Футбольный бокс, заказ {order_id}",
    "return": "https://1lyas1k333.github.io",          # ← заменил
    "failedUrl": "https://1lyas1k333.github.io",       # ← заменил
    "payload": order_id,
    "callback_url": "https://arturchik-box-2.onrender.com/platega-webhook"
}
    
    try:
        response = requests.post(PLATEGA_API_URL, json=payload, headers=headers, timeout=30)
        result = response.json()
        if response.status_code == 200 and result.get('redirect'):
            return jsonify({"success": True, "order_id": order_id, "amount": amount, "payment_url": result['redirect']})
        else:
            return jsonify({"success": False, "error": "Ошибка создания платежа"}), 500
    except Exception as e:
        print(f"[PLATEGA] Ошибка: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/platega-webhook', methods=['POST', 'GET'])
def platega_webhook():
    if request.method == 'GET':
        return jsonify({"status": "ok", "message": "Webhook is active"}), 200
    data = request.get_json()
    print(f"[WEBHOOK] {data}")
    order_id = data.get('payload') or data.get('order_id')
    if order_id and data.get('status') == 'CONFIRMED':
        update_order_status(order_id, 'paid')
        send_telegram_message(f"✅ ОПЛАЧЕН ЗАКАЗ\n📦 Заказ: {order_id}")
    return jsonify({"success": True}), 200
@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        data = request.get_json()
        print(f"[WEBHOOK] Получены данные: {data}")
        
        message = data.get('message', {})
        chat_id = message.get('chat', {}).get('id')
        text = message.get('text', '')
        
        print(f"[WEBHOOK] chat_id={chat_id}, text={text}")
        
        if not chat_id:
            print("[WEBHOOK] Нет chat_id")
            return jsonify({"ok": False}), 400
        
        if str(chat_id) == TELEGRAM_CHAT_ID:
            print("[WEBHOOK] Сообщение от админа, игнорируем")
            return jsonify({"ok": True}), 200
        
        if text and text.startswith('/status'):
            print("[WEBHOOK] Обнаружена команда /status")
            parts = text.split(' ')
            if len(parts) > 1:
                order_id = parts[1]
                print(f"[WEBHOOK] Ищем заказ: {order_id}")
                res = supabase.table("orders").select("*").eq("order_id", order_id).execute()
                if res.data:
                    order = res.data[0]
                    status_text = {
                        'pending': '⏳ Ожидает оплаты',
                        'paid': '✅ Оплачен',
                        'shipped': '📦 Отправлен',
                        'completed': '🎉 Завершён'
                    }.get(order['status'], order['status'])
                    
                    msg = f"""📦 <b>Заказ {order_id}</b>

👤 Получатель: {order.get('customer_name', '—')}
💰 Сумма: {order.get('total_amount', 0)} ₽
📌 Статус: {status_text}
📅 Дата: {order.get('created_at', '—')[:10]}"""
                    
                    if order.get('tracking_number'):
                        msg += f"\n\n📦 Трек-номер: {order['tracking_number']}\n🔗 Отследить: https://www.cdek.ru/track?order_id={order['tracking_number']}"
                    
                    print(f"[WEBHOOK] Отправляем ответ пользователю {chat_id}")
                    send_telegram_to_user(chat_id, msg)
                else:
                    print(f"[WEBHOOK] Заказ {order_id} не найден")
                    send_telegram_to_user(chat_id, "❌ Заказ не найден. Проверьте номер.")
            else:
                print("[WEBHOOK] Неверный формат команды")
                send_telegram_to_user(chat_id, "ℹ️ Используйте: /status НОМЕР_ЗАКАЗА\nНапример: /status ORDER_20260611123456")
        else:
            print("[WEBHOOK] Неизвестная команда")
            send_telegram_to_user(chat_id, "ℹ️ Доступные команды:\n/status НОМЕР — узнать статус заказа")
        
        return jsonify({"ok": True}), 200
        
    except Exception as e:
        print(f"[WEBHOOK] Ошибка: {e}")
        return jsonify({"ok": False}), 500
@app.route('/api/set-telegram', methods=['POST'])
def set_telegram():
    try:
        data = request.get_json()
        telegram_id = data.get('telegram_id')
        email = data.get('email')
        
        print(f"[DEBUG] set_telegram: telegram_id={telegram_id}, email={email}")
        
        if not telegram_id:
            return jsonify({'success': False, 'error': 'Telegram ID не указан'}), 400
        
        if not email:
            return jsonify({'success': False, 'error': 'Email не указан'}), 400
        
        # Обновляем telegram_id у пользователя
        res = supabase.table("users").update({"telegram_id": telegram_id}).eq("email", email).execute()
        
        if not res.data:
            return jsonify({'success': False, 'error': 'Пользователь с таким email не найден'}), 404
        
        # Отправляем тестовое сообщение
        try:
            send_telegram_to_user(telegram_id, "✅ Ваш Telegram успешно привязан к аккаунту АРТУРЧИК box!")
        except Exception as e:
            print(f"[TG] Ошибка отправки: {e}")
        
        return jsonify({'success': True, 'message': 'Telegram успешно привязан!'})
        
    except Exception as e:
        print(f"[ERROR] set_telegram: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# === АДМИН-ПАНЕЛЬ ===
ADMIN_HTML = '''
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Админ-панель | АРТУРЧИК box</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: system-ui;
            background: linear-gradient(135deg, #0a0a0a 0%, #1a2a1a 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { max-width: 1400px; margin: 0 auto; }
        h1 { color: #2d8c4e; margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px; }
        .stats { display: flex; gap: 20px; margin-bottom: 30px; flex-wrap: wrap; }
        .stat-card { background: rgba(0,0,0,0.6); backdrop-filter: blur(10px); padding: 20px; border-radius: 20px; border: 1px solid #2d8c4e; min-width: 150px; text-align: center; }
        .stat-number { font-size: 36px; font-weight: bold; color: #2d8c4e; }
        .stat-label { color: #ccc; font-size: 14px; }
        .export-btn { background: #2d8c4e; color: white; border: none; padding: 12px 24px; border-radius: 12px; cursor: pointer; font-size: 16px; margin-bottom: 20px; }
        table { width: 100%; background: rgba(0,0,0,0.6); border-radius: 20px; overflow: hidden; border-collapse: collapse; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid rgba(255,255,255,0.1); color: white; }
        th { background: #2d8c4e; cursor: pointer; }
        tr:hover { background: rgba(45,140,78,0.2); }
        select { background: #1a2a1a; color: white; border: 1px solid #2d8c4e; padding: 5px 10px; border-radius: 8px; cursor: pointer; }
        .refresh-btn, .logout-btn { background: #2d8c4e; color: white; border: none; padding: 8px 16px; border-radius: 8px; cursor: pointer; }
        .logout-btn { background: #ff4444; }
        @media (max-width: 768px) { th, td { padding: 8px; font-size: 12px; } .stats { gap: 10px; } .stat-card { padding: 10px; min-width: 100px; } .stat-number { font-size: 24px; } }
    </style>
</head>
<body>
    <div class="container">
        <h1>📦 Админ-панель АРТУРЧИК box <div><button class="refresh-btn" onclick="loadOrders()">🔄 Обновить</button><button class="logout-btn" onclick="logout()">🚪 Выйти</button></div></h1>
        
        <div class="stats">
            <div class="stat-card"><div class="stat-number" id="totalOrders">0</div><div class="stat-label">Всего заказов</div></div>
            <div class="stat-card"><div class="stat-number" id="totalAmount">0</div><div class="stat-label">Сумма (₽)</div></div>
            <div class="stat-card"><div class="stat-number" id="pendingOrders">0</div><div class="stat-label">Ожидают оплаты</div></div>
        </div>
        
        <div style="display: flex; gap: 15px; margin-bottom: 20px; flex-wrap: wrap; align-items: center;">
            <select id="statusFilter" onchange="applyFilters()" style="background: #1a2a1a; color: white; border: 1px solid #2d8c4e; padding: 10px 15px; border-radius: 10px;">
                <option value="all">📋 Все статусы</option>
                <option value="pending">⏳ Ожидают оплаты</option>
                <option value="paid">✅ Оплаченные</option>
                <option value="shipped">📦 Отправленные</option>
                <option value="completed">🎉 Завершённые</option>
            </select>
            
            <input type="date" id="dateFrom" onchange="applyFilters()" style="background: #1a2a1a; color: white; border: 1px solid #2d8c4e; padding: 10px 15px; border-radius: 10px;">
            <span style="color: white;">—</span>
            <input type="date" id="dateTo" onchange="applyFilters()" style="background: #1a2a1a; color: white; border: 1px solid #2d8c4e; padding: 10px 15px; border-radius: 10px;">
            
            <button onclick="resetFilters()" style="background: #2d8c4e; color: white; border: none; padding: 10px 20px; border-radius: 10px; cursor: pointer;">🔄 Сбросить</button>
        </div>
        
        <button class="export-btn" onclick="exportExcel()">📊 Экспорт в Excel</button>
        
        <table>
            <thead>
                <tr>
                    <th>№ заказа</th>
                    <th>Покупатель</th>
                    <th>Email/Телефон</th>
                    <th>Сумма</th>
                    <th>Статус</th>
                    <th>Трек-номер</th>
                    <th>Дата</th>
                    <th></th>
                </tr>
            </thead>
            <tbody id="ordersBody">
                <tr><td colspan="8">Загрузка...</td></tr>
            </tbody>
        </table>
    </div>
    <script>
        let ordersData = [];
        let allOrdersData = [];
        
        function applyFilters() {
            const status = document.getElementById('statusFilter').value;
            const dateFrom = document.getElementById('dateFrom').value;
            const dateTo = document.getElementById('dateTo').value;
            let filtered = [...allOrdersData];
            if (status !== 'all') {
                filtered = filtered.filter(order => order.status === status);
            }
            if (dateFrom) {
                const fromDate = new Date(dateFrom);
                filtered = filtered.filter(order => new Date(order.created_at) >= fromDate);
            }
            if (dateTo) {
                const toDate = new Date(dateTo);
                toDate.setHours(23, 59, 59);
                filtered = filtered.filter(order => new Date(order.created_at) <= toDate);
            }
            ordersData = filtered;
            renderOrders();
            updateStats();
        }
        
        function resetFilters() {
            document.getElementById('statusFilter').value = 'all';
            document.getElementById('dateFrom').value = '';
            document.getElementById('dateTo').value = '';
            ordersData = [...allOrdersData];
            renderOrders();
            updateStats();
        }
        
        async function loadOrders() {
            try {
                const response = await fetch('/api/orders');
                const data = await response.json();
                if (data.success) { 
                    allOrdersData = data.orders; 
                    ordersData = [...allOrdersData];
                    renderOrders(); 
                    updateStats(); 
                }
            } catch(error) { 
                document.getElementById('ordersBody').innerHTML = '<tr><td colspan="8">Ошибка загрузки</td></tr>'; 
            }
        }
        
        function renderOrders() {
            const tbody = document.getElementById('ordersBody');
            if (!ordersData.length) { 
                tbody.innerHTML = '<tr><td colspan="8">Нет заказов</td></tr>'; 
                return; 
            }
            tbody.innerHTML = ordersData.map(order => `
                <tr>
                    <td>${order.order_id}</td>
                    <td>${order.customer_name || '—'}</td>
                    <td>${order.customer_email || '—'}<br><small>${order.customer_phone || ''}</small></td>
                    <td>${order.total_amount} ₽</td>
                    <td>
                        <select onchange="updateStatus('${order.order_id}', this.value)">
                            <option value="pending" ${order.status === 'pending' ? 'selected' : ''}>⏳ Ожидает</option>
                            <option value="paid" ${order.status === 'paid' ? 'selected' : ''}>✅ Оплачен</option>
                            <option value="shipped" ${order.status === 'shipped' ? 'selected' : ''}>📦 Отправлен</option>
                            <option value="completed" ${order.status === 'completed' ? 'selected' : ''}>🎉 Завершён</option>
                        </select>
                    </td>
                    <td>
                        <input type="text" id="track_${order.order_id}" 
                               value="${order.tracking_number || ''}" 
                               placeholder="Введите трек-номер"
                               onchange="updateTracking('${order.order_id}', this.value)"
                               style="background: #1a2a1a; color: white; border: 1px solid #2d8c4e; padding: 5px; border-radius: 8px; width: 150px;">
                    </td>
                    <td>${new Date(order.created_at).toLocaleString()}</td>
                    <td>
                        <button onclick="showOrderDetails('${order.order_id}')" 
                                style="background: #2d8c4e; border: none; padding: 5px 10px; border-radius: 5px; cursor: pointer; color: white;">
                            📋 Подробнее
                        </button>
                    </td>
                </tr>
            `).join('');
        }
        
        function updateStats() {
            document.getElementById('totalOrders').textContent = ordersData.length;
            document.getElementById('totalAmount').textContent = ordersData.reduce((s,o) => s + o.total_amount, 0).toLocaleString();
            document.getElementById('pendingOrders').textContent = ordersData.filter(o => o.status === 'pending').length;
        }
        
        async function updateStatus(orderId, newStatus) {
            await fetch('/api/update-status', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({order_id: orderId, status: newStatus})});
            loadOrders();
        }
        
      async function updateTracking(orderId, trackingNumber) {
    try {
        const response = await fetch('/api/update-tracking', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ order_id: orderId, tracking_number: trackingNumber })
        });
        const data = await response.json();
        if (data.success) {
            showToast('✅ Трек-номер сохранён, уведомление отправлено клиенту');
        } else {
            showToast('❌ Ошибка: ' + data.error);
        }
    } catch (error) {
        console.error('Ошибка:', error);
        showToast('❌ Ошибка соединения');
    }
}
        
        async function showOrderDetails(orderId) {
            try {
                const response = await fetch('/api/orders');
                const data = await response.json();
                if (!data.success) {
                    alert('Ошибка загрузки заказа');
                    return;
                }
                const order = data.orders.find(o => o.order_id === orderId);
                if (!order) {
                    alert('Заказ не найден');
                    return;
                }
                
                let items = [];
                try {
                    items = JSON.parse(order.items || '[]');
                } catch(e) { items = []; }
                
                const itemsHtml = items.map(item => `
                    <div style="margin-bottom: 10px; padding: 8px; background: rgba(255,255,255,0.1); border-radius: 8px; color: white;">
                        <strong style="color: #2d8c4e;">${item.name}</strong><br>
                        <span>📏 Размер: ${item.size}</span><br>
                        <span>🚫 Исключения: ${item.exclusions?.join(', ') || 'нет'}</span><br>
                        <span>📦 Количество: ${item.quantity}</span><br>
                        <span>💰 Цена: ${item.price} ₽</span>
                    </div>
                `).join('');
                
                const modalHtml = `
                    <div id="orderDetailModal" style="position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%); background: #0d1f0d; padding: 25px; border-radius: 20px; z-index: 10003; max-width: 500px; width: 90%; max-height: 80vh; overflow-y: auto; border: 1px solid #2d8c4e;">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                            <h3 style="color: #2d8c4e;">📦 Заказ ${order.order_id}</h3>
                            <button onclick="closeOrderDetailModal()" style="background: #ff4444; border: none; padding: 5px 10px; border-radius: 5px; cursor: pointer; color: white;">✕</button>
                        </div>
                        <div style="margin-bottom: 10px;"><strong style="color: #2d8c4e;">👤 Покупатель:</strong> <span style="color: white;">${order.customer_name || '—'}</span></div>
                        <div style="margin-bottom: 10px;"><strong style="color: #2d8c4e;">📧 Email:</strong> <span style="color: white;">${order.customer_email || '—'}</span></div>
                        <div style="margin-bottom: 10px;"><strong style="color: #2d8c4e;">📞 Телефон:</strong> <span style="color: white;">${order.customer_phone || '—'}</span></div>
                        <div style="margin-bottom: 10px;"><strong style="color: #2d8c4e;">📍 Адрес:</strong> <span style="color: white;">${order.customer_city || '—'}, ${order.customer_address || '—'}</span></div>
                        <div style="margin-bottom: 10px;"><strong style="color: #2d8c4e;">📋 Состав заказа:</strong></div>
                        <div>${itemsHtml}</div>
                        <div style="margin-top: 10px;"><strong style="color: #2d8c4e;">💰 Сумма:</strong> <span style="color: white;">${order.total_amount} ₽</span></div>
                        <div><strong style="color: #2d8c4e;">📌 Статус:</strong> <span style="color: white;">${order.status === 'pending' ? '⏳ Ожидает оплаты' : order.status === 'paid' ? '✅ Оплачен' : order.status === 'shipped' ? '📦 Отправлен' : '🎉 Завершён'}</span></div>
                        <div><strong style="color: #2d8c4e;">📅 Дата:</strong> <span style="color: white;">${new Date(order.created_at).toLocaleString()}</span></div>
                        ${order.customer_notes ? `<div><strong style="color: #2d8c4e;">📝 Примечание:</strong> <span style="color: white;">${order.customer_notes}</span></div>` : ''}
                    </div>
                    <div id="orderDetailOverlay" onclick="closeOrderDetailModal()" style="position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 10002;"></div>
                `;
                document.body.insertAdjacentHTML('beforeend', modalHtml);
            } catch (error) {
                console.error('Ошибка:', error);
                alert('Не удалось загрузить данные заказа');
            }
        }
        
        function closeOrderDetailModal() {
            const modal = document.getElementById('orderDetailModal');
            const overlay = document.getElementById('orderDetailOverlay');
            if (modal) modal.remove();
            if (overlay) overlay.remove();
        }

        function showToast(message) {
    const toast = document.createElement('div');
    toast.textContent = message;
    toast.style.position = 'fixed';
    toast.style.bottom = '20px';
    toast.style.left = '50%';
    toast.style.transform = 'translateX(-50%)';
    toast.style.backgroundColor = '#2d8c4e';
    toast.style.color = 'white';
    toast.style.padding = '12px 24px';
    toast.style.borderRadius = '50px';
    toast.style.zIndex = '10000';
    toast.style.fontSize = '14px';
    document.body.appendChild(toast);
    setTimeout(() => toast.remove(), 2500);
}
        
        function exportExcel() { window.open('/export-orders', '_blank'); }
        function logout() { window.location.href = '/admin/logout'; }
        
        loadOrders();
        setInterval(loadOrders, 30000);
    </script>
</body>
</html>
'''

@app.route('/admin', methods=['GET', 'POST'])
def admin_panel():
    if request.method == 'POST':
        if request.form.get('password') == '123':
            return ADMIN_HTML
        return '<h2>❌ Неверный пароль</h2><a href="/admin">Назад</a>'
    return '''
    <!DOCTYPE html>
    <html>
    <head><meta charset="UTF-8"><title>Вход в админ-панель</title>
    <style>
        body {
            font-family: system-ui;
            background: linear-gradient(135deg, #0a0a0a 0%, #1a2a1a 100%);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            margin: 0;
        }
        .login-box {
            background: rgba(0,0,0,0.6);
            padding: 40px;
            border-radius: 30px;
            text-align: center;
            border: 1px solid #2d8c4e;
        }
        h2 { color: #2d8c4e; margin-bottom: 20px; }
        input { padding: 12px; width: 200px; border-radius: 10px; border: 1px solid #2d8c4e; background: rgba(0,0,0,0.5); color: white; }
        button { margin-top: 15px; padding: 10px 20px; background: #2d8c4e; border: none; border-radius: 10px; color: white; cursor: pointer; }
    </style>
    </head>
    <body>
        <div class="login-box">
            <h2>🔐 Вход в админ-панель</h2>
            <form method="post">
                <input type="password" name="password" placeholder="Введите пароль" autofocus>
                <br><button type="submit">Войти</button>
            </form>
        </div>
    </body>
    </html>
    '''

@app.route('/admin/logout')
def admin_logout():
    return redirect('/admin')

@app.route('/export-orders', methods=['GET'])
def export_orders():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        orders = supabase.table("orders").select("*").execute()
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        headers = ['№ заказа', 'Покупатель', 'Email', 'Телефон', 'Товары', 'Сумма (₽)', 'Статус', 'Дата создания']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2d8c4e", end_color="2d8c4e", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, order in enumerate(orders.data, 2):
            items_text = ''
            try:
                items = json.loads(order.get('items', '[]'))
                items_text = ', '.join([f"{item.get('name', '')} ({item.get('size', '')}) x{item.get('quantity', 1)}" for item in items])
            except:
                pass
            ws.cell(row=row_idx, column=1, value=order.get('order_id', ''))
            ws.cell(row=row_idx, column=2, value=order.get('customer_name', ''))
            ws.cell(row=row_idx, column=3, value=order.get('customer_email', ''))
            ws.cell(row=row_idx, column=4, value=order.get('customer_phone', ''))
            ws.cell(row=row_idx, column=5, value=items_text)
            ws.cell(row=row_idx, column=6, value=order.get('total_amount', 0))
            ws.cell(row=row_idx, column=7, value=order.get('status', 'pending'))
            ws.cell(row=row_idx, column=8, value=order.get('created_at', ''))
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 20
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/')
def home():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.getenv("PORT", 10000)))
